import requests
from bs4 import BeautifulSoup
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

def scrape_caf_champions_league():
    """
    Scrape CAF Champions League 2024-2025 season matches from Transfermarkt
    """
    
    # Base URL for match day view
    base_url = "https://www.transfermarkt.com/caf-champions-league/spieltag/pokalwettbewerb/ACL/plus/0"
    
    # List of rounds to scrape based on the dropdown options
    rounds_to_scrape = [
        {'name': 'First Round 1st leg', 'spieltag': '1'},
        {'name': 'First Round 2nd leg', 'spieltag': '2'},
        {'name': 'Second Round 1st leg', 'spieltag': '3'},
        {'name': 'Second Round 2nd leg', 'spieltag': '4'},
        # Note: Group stage uses same spieltag numbers but different context
        # We'll handle this by checking if matches exist
    ]
    
    # Add group stage matchdays
    for i in range(1, 7):
        rounds_to_scrape.append({
            'name': f'Group Stage - Matchday {i}',
            'spieltag': str(i),
            'is_group': True
        })
    
    # Add knockout rounds
    knockout_rounds = [
        {'name': 'Quarter-Finals 1st leg', 'spieltag': '11'},
        {'name': 'Quarter-Finals 2nd leg', 'spieltag': '12'},
        {'name': 'Semi-Finals 1st leg', 'spieltag': '13'},
        {'name': 'Semi-Finals 2nd leg', 'spieltag': '14'},
        {'name': 'Final 1st leg', 'spieltag': '15'},
        {'name': 'Final 2nd leg', 'spieltag': '16'},
    ]
    rounds_to_scrape.extend(knockout_rounds)
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    }
    
    all_matches = []
    scraped_rounds = set()
    
    for round_info in rounds_to_scrape:
        round_name = round_info['name']
        spieltag = round_info['spieltag']
        
        # Skip if we already scraped this spieltag (for group stage overlap)
        round_key = f"{spieltag}_{round_info.get('is_group', False)}"
        if round_key in scraped_rounds and round_info.get('is_group', False):
            continue
        
        url = f"{base_url}?saison_id=2024&spieltag={spieltag}"
        
        print(f"\n{'='*60}")
        print(f"Scraping: {round_name}")
        print(f"URL: {url}")
        print(f"{'='*60}")
        
        try:
            # Add delay to be respectful
            time.sleep(1)
            
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find all match tables
            tables = soup.find_all('table')
            
            match_count = 0
            
            for table in tables:
                # Skip the filter table
                if 'auflistung' in table.get('class', []):
                    continue
                
                rows = table.find_all('tr')
                
                # A match table typically has at least 3 rows
                if len(rows) >= 3:
                    try:
                        # Row 1: Teams and score
                        row1_cells = rows[0].find_all('td')
                        if len(row1_cells) >= 5:
                            # Extract home team (first cell)
                            home_team = row1_cells[0].get_text(strip=True)
                            
                            # Extract score (middle cells)
                            score = ""
                            for cell in row1_cells[2:6]:
                                score_link = cell.find('a', href=lambda x: x and '/spielbericht/' in x)
                                if score_link:
                                    score_text = score_link.get_text(strip=True)
                                    # Clean up score
                                    score_match = re.search(r'\d+:\d+', score_text)
                                    score = score_match.group() if score_match else score_text
                                    break
                                # Also check for text scores
                                elif ':' in cell.get_text():
                                    score_match = re.search(r'\d+:\d+', cell.get_text())
                                    if score_match:
                                        score = score_match.group()
                                        break
                            
                            # Extract away team (last relevant cell)
                            away_team = ""
                            for cell in reversed(row1_cells[-3:]):
                                text = cell.get_text(strip=True)
                                if text and text not in ['', score] and len(text) > 2:
                                    # Make sure it's not a score or abbreviation only
                                    if not re.match(r'^\d+:\d+$', text):
                                        away_team = text
                                        break
                            
                            # Row 2: First leg info (if exists)
                            first_leg = ""
                            if len(rows) > 1:
                                row2_text = rows[1].get_text(strip=True)
                                if 'First leg' in row2_text or 'first leg' in row2_text:
                                    first_leg = row2_text
                            
                            # Find date/time in subsequent rows
                            date = ""
                            time_str = ""
                            for row in rows[1:]:
                                row_text = row.get_text(strip=True)
                                # Look for date pattern
                                date_match = re.search(r'(\w+,\w+\d{2}/\d{2}/\d{4})', row_text)
                                if date_match:
                                    date = date_match.group(1)
                                    # Clean up date format
                                    date = re.sub(r'(\w+),(\w+)', r'\2 ', date)
                                
                                # Look for time pattern
                                time_match = re.search(r'(\d{1,2}:\d{2}\s*[AP]M)', row_text, re.IGNORECASE)
                                if time_match:
                                    time_str = time_match.group(1)
                                
                                if date and time_str:
                                    break
                            
                            # Only add if we have valid team names
                            if home_team and away_team and home_team != away_team and len(home_team) > 2 and len(away_team) > 2:
                                match = {
                                    "round": round_name,
                                    "date": date,
                                    "time": time_str,
                                    "home_team": home_team,
                                    "away_team": away_team,
                                    "score": score if score else "TBD",
                                    "first_leg": first_leg
                                }
                                all_matches.append(match)
                                match_count += 1
                                print(f"  {home_team} vs {away_team} - {score if score else 'TBD'} ({date})")
                    
                    except Exception as e:
                        # Skip malformed tables
                        continue
            
            if match_count > 0:
                print(f"\n✓ Found {match_count} matches in {round_name}")
                scraped_rounds.add(round_key)
            else:
                print(f"\n  No matches found (round may not have started yet)")
            
        except Exception as e:
            print(f"Error scraping {round_name}: {e}")
            continue
    
    return all_matches

def create_excel(matches, excel_file):
    """
    Create formatted Excel file from matches
    """
    print(f"\n{'='*60}")
    print("Creating Excel file...")
    print(f"{'='*60}\n")
    
    # Convert to DataFrame
    df = pd.DataFrame(matches)
    
    # Reorder columns
    column_order = ['round', 'date', 'time', 'home_team', 'away_team', 'score', 'first_leg']
    df = df[column_order]
    
    # Rename columns
    df.columns = ['Round', 'Date', 'Time', 'Home Team', 'Away Team', 'Score', 'First Leg Info']
    
    # Write to Excel
    df.to_excel(excel_file, index=False, sheet_name='CAF Champions League')
    
    # Format the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    round_fill = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
    round_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Track current round for alternating colors
    current_round = None
    round_color = True
    
    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        round_value = row[0].value
        if round_value != current_round:
            current_round = round_value
            round_color = not round_color
        
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if round_color:
                cell.fill = round_fill
        
        row[0].font = round_font
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save
    wb.save(excel_file)
    print(f"✓ Excel file saved: {excel_file}")

def main():
    import os
    
    print("="*60)
    print("CAF Champions League 2024-2025 - Complete Scraper")
    print("="*60 + "\n")
    
    # Get script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_file = os.path.join(script_dir, "caf_champions_league_2024_2025.json")
    excel_file = os.path.join(script_dir, "CAF_Champions_League_2024_2025.xlsx")
    
    # Scrape data from ALL rounds
    matches = scrape_caf_champions_league()
    
    if not matches:
        print("\n❌ No matches found!")
        return
    
    print(f"\n{'='*60}")
    print(f"TOTAL MATCHES SCRAPED: {len(matches)}")
    print(f"{'='*60}\n")
    
    # Save JSON
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(matches, f, indent=2, ensure_ascii=False)
    print(f"✓ JSON saved: {json_file}")
    
    # Create Excel
    try:
        create_excel(matches, excel_file)
        
        print(f"\n{'='*60}")
        print("SUCCESS!")
        print(f"{'='*60}")
        print(f"Excel file: {excel_file}")
        print(f"JSON file: {json_file}")
        print(f"Total matches: {len(matches)}")
        
        # Print summary by round
        print(f"\nMatches by Round:")
        print("-" * 60)
        rounds = {}
        for match in matches:
            round_name = match['round']
            if round_name not in rounds:
                rounds[round_name] = 0
            rounds[round_name] += 1
        
        for round_name, count in rounds.items():
            print(f"{round_name}: {count} matches")
        
        print(f"{'='*60}\n")
        
    except PermissionError:
        print(f"\n❌ Error: Cannot write to {excel_file}")
        print("The file might be open in Excel. Please close it and run again.\n")
    except Exception as e:
        print(f"\n❌ Error creating Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()