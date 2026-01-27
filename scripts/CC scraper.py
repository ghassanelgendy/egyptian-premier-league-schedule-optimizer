import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import time
import re
import json
import os

PROGRESS_FILE = 'scraping_progress.json'
DATA_FILE = 'scraped_matches.json'

def load_progress():
    """Load scraping progress from file"""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    return {'completed_rounds': [], 'last_round_index': -1}

def save_progress(progress):
    """Save scraping progress to file"""
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, indent=2)

def load_matches():
    """Load previously scraped matches"""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_matches(matches):
    """Save matches to file"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(matches, f, indent=2, ensure_ascii=False)

def get_all_rounds(base_url, headers):
    """Detect all available rounds from the page"""
    try:
        response = requests.get(base_url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        select = soup.find('select', {'name': 'gruppe'})
        
        if select:
            rounds = []
            for option in select.find_all('option'):
                round_code = option.get('value', '')
                round_name = option.get_text(strip=True)
                if round_code and round_name:
                    rounds.append((round_code, round_name))
            return rounds
    except Exception as e:
        print(f"Error detecting rounds: {e}")
    
    return []

def scrape_transfermarkt_caf_cup_resumable():
    """
    Scrape CAF Confederation Cup 2024 with resume capability
    """
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://www.transfermarkt.com/'
    }
    
    base_url = 'https://www.transfermarkt.com/caf-confederation-cup/spieltag/pokalwettbewerb/CAFC/plus/0?saison_id=2024&gruppe=1RH'
    
    # Load progress and existing matches
    progress = load_progress()
    all_matches = load_matches()
    
    if progress['last_round_index'] >= 0:
        print("\n" + "=" * 60)
        print("⚠ RESUMING FROM PREVIOUS SESSION")
        print("=" * 60)
        print(f"Previously scraped: {len(all_matches)} matches")
        print(f"Completed rounds: {len(progress['completed_rounds'])}")
        print("=" * 60 + "\n")
    
    print("Detecting available rounds...")
    rounds = get_all_rounds(base_url, headers)
    
    if not rounds:
        print("Could not auto-detect rounds. Using fallback list.")
        rounds = [
            ('1RH', 'First Round 1st leg'),
            ('1RR', 'First Round 2nd leg'),
            ('2RH', 'Second Round 1st leg'),
            ('2RR', 'Second Round 2nd leg'),
            ('A', 'Group A'),
            ('B', 'Group B'),
            ('C', 'Group C'),
            ('D', 'Group D'),
            ('VFH', 'Quarter-Finals 1st leg'),
            ('VFR', 'Quarter-Finals 2nd leg'),
            ('HFH', 'Semi-Finals 1st Leg'),
            ('HFR', 'Semi-Finals 2nd Leg'),
            ('FFH', 'Final 1st leg'),
            ('FFR', 'Final 2nd leg')
        ]
    
    print(f"Found {len(rounds)} rounds total\n")
    
    print("Starting to scrape CAF Confederation Cup 2024 data...")
    print("=" * 60)
    
    for idx, (round_code, round_name) in enumerate(rounds):
        # Skip already completed rounds
        if round_code in progress['completed_rounds']:
            print(f"\n✓ Skipping {round_name} (already completed)")
            continue
        
        url = f'https://www.transfermarkt.com/caf-confederation-cup/spieltag/pokalwettbewerb/CAFC/plus/0?saison_id=2024&gruppe={round_code}'
        
        print(f"\n[{idx+1}/{len(rounds)}] Scraping {round_name}...")
        
        try:
            # Retry logic for failed requests
            max_retries = 3
            retry_count = 0
            response = None
            
            while retry_count < max_retries:
                try:
                    response = requests.get(url, headers=headers, timeout=20)
                    response.raise_for_status()
                    break
                except requests.exceptions.RequestException as e:
                    retry_count += 1
                    if retry_count < max_retries:
                        wait_time = retry_count * 10  # 10, 20, 30 seconds
                        print(f"  ⚠ Connection failed, retrying in {wait_time}s... (attempt {retry_count}/{max_retries})")
                        time.sleep(wait_time)
                    else:
                        print(f"  ✗ Failed after {max_retries} attempts: {str(e)[:100]}")
                        print(f"\n{'='*60}")
                        print("⚠ SCRAPING PAUSED - CONNECTION ISSUE")
                        print("="*60)
                        print(f"Progress saved: {len(all_matches)} matches scraped so far")
                        print(f"Completed: {len(progress['completed_rounds'])}/{len(rounds)} rounds")
                        print("\nTo resume:")
                        print("1. Check your internet connection")
                        print("2. Wait a few minutes if blocked")
                        print("3. Run the script again - it will continue from here")
                        print("="*60)
                        
                        # Save current progress before exiting
                        save_progress(progress)
                        save_matches(all_matches)
                        return None
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            matches_found = 0
            round_matches = []
            
            boxes = soup.find_all('div', class_='box')
            
            for box in boxes:
                # Extract date from box
                box_date = ''
                box_text = box.get_text()
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', box_text)
                if date_match:
                    box_date = date_match.group(1)
                
                table = box.find('table')
                if not table:
                    continue
                
                rows = table.find_all('tr')
                
                for row in rows:
                    try:
                        cells = row.find_all('td')
                        
                        if len(cells) < 8:
                            continue
                        
                        # Check for result cell
                        result_cell = None
                        for cell in cells:
                            if 'spieltagsansicht-ergebnis' in cell.get('class', []):
                                result_cell = cell
                                break
                        
                        if not result_cell:
                            continue
                        
                        # Extract teams
                        home_team = ''
                        away_team = ''
                        
                        team_cells = []
                        for cell in cells:
                            cell_classes = cell.get('class', [])
                            if 'spieltagsansicht-vereinsname' in cell_classes and 'hide-for-small' in cell_classes:
                                team_cells.append(cell)
                        
                        if len(team_cells) >= 2:
                            home_cell = team_cells[0]
                            away_cell = team_cells[1]
                            
                            home_link = home_cell.find('a', href=lambda x: x and '/verein/' in x)
                            away_link = away_cell.find('a', href=lambda x: x and '/verein/' in x)
                            
                            if home_link:
                                home_team = home_link.get('title', home_link.get_text(strip=True))
                            if away_link:
                                away_team = away_link.get('title', away_link.get_text(strip=True))
                        
                        if not home_team or not away_team:
                            continue
                        
                        # Extract result
                        result = result_cell.get_text(strip=True)
                        
                        if not result or result == '-:-':
                            result = 'Not Played'
                        elif not re.match(r'^\d+:\d+$', result):
                            result = 'Not Played'
                        
                        date = box_date if box_date else 'TBD'
                        
                        match_data = {
                            'Round': round_name,
                            'Date': date,
                            'Home Team': home_team,
                            'Away Team': away_team,
                            'Result': result
                        }
                        
                        round_matches.append(match_data)
                        matches_found += 1
                    
                    except Exception as e:
                        continue
            
            # Add this round's matches to all matches
            all_matches.extend(round_matches)
            
            # Mark round as completed
            progress['completed_rounds'].append(round_code)
            progress['last_round_index'] = idx
            
            # Save progress after each successful round
            save_progress(progress)
            save_matches(all_matches)
            
            print(f"  ✓ Found {matches_found} matches")
            print(f"  ✓ Progress saved ({len(all_matches)} total matches)")
            
            # Longer delay between rounds
            if idx < len(rounds) - 1:  # Don't wait after last round
                print(f"  ⏳ Waiting 5 seconds before next round...")
                time.sleep(5)
            
        except Exception as e:
            print(f"  ✗ Unexpected error: {e}")
            save_progress(progress)
            save_matches(all_matches)
            continue
    
    print("\n" + "=" * 60)
    print(f"Total matches scraped: {len(all_matches)}")
    
    if all_matches:
        # Create DataFrame
        df = pd.DataFrame(all_matches)
        
        df['Home Team'] = df['Home Team'].str.strip()
        df['Away Team'] = df['Away Team'].str.strip()
        
        df.insert(0, 'No.', range(1, len(df) + 1))
        
        # Export to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'CAF_Confederation_Cup_2024_COMPLETE_{timestamp}.xlsx'
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='CAF Confederation Cup 2024')
            
            worksheet = writer.sheets['CAF Confederation Cup 2024']
            
            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 28
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 35
            worksheet.column_dimensions['E'].width = 35
            worksheet.column_dimensions['F'].width = 12
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                          min_col=1, max_col=6):
                for cell in row:
                    cell.border = border
                    if cell.column in [1, 3, 6]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center')
        
        print(f"\n✓ Data exported successfully to: {filename}")
        print("\n" + "=" * 60)
        print("Preview of data (first 25 matches):")
        print("=" * 60)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 35)
        print(df.head(25).to_string(index=False))
        
        print("\n" + "=" * 60)
        print("Summary by Round:")
        print("=" * 60)
        summary = df.groupby('Round', sort=False).size()
        for round_name, count in summary.items():
            print(f"  {round_name:.<30} {count:>3} matches")
        
        print("\n" + "=" * 60)
        print("Statistics:")
        print("=" * 60)
        played = df[df['Result'] != 'Not Played'].shape[0]
        not_played = df[df['Result'] == 'Not Played'].shape[0]
        print(f"  Total matches:............ {len(df)}")
        print(f"  Matches played:........... {played}")
        print(f"  Matches not played:....... {not_played}")
        
        all_teams = sorted(set(df['Home Team'].tolist() + df['Away Team'].tolist()))
        print(f"  Unique teams:............. {len(all_teams)}")
        
        print("\n" + "=" * 60)
        print("Teams participating:")
        print("=" * 60)
        for i, team in enumerate(all_teams, 1):
            print(f"  {i:2d}. {team}")
        
        # Clean up progress files after successful completion
        if len(progress['completed_rounds']) == len(rounds):
            if os.path.exists(PROGRESS_FILE):
                os.remove(PROGRESS_FILE)
            if os.path.exists(DATA_FILE):
                os.remove(DATA_FILE)
            print("\n✓ Progress files cleaned up (scraping complete)")
        
        return df
    else:
        print("\n✗ No matches found.")
        return None

if __name__ == "__main__":
    try:
        print("\n" + "=" * 60)
        print("CAF CONFEDERATION CUP 2024 - RESUMABLE SCRAPER")
        print("=" * 60)
        print("\nFeatures:")
        print("  ✓ Auto-saves progress after each round")
        print("  ✓ Auto-resumes if interrupted")
        print("  ✓ 3 retry attempts with delays")
        print("  ✓ 5 second delays between rounds")
        print("\nFiles created:")
        print("  - scraping_progress.json (progress tracking)")
        print("  - scraped_matches.json (temporary data)")
        print("  - CAF_Confederation_Cup_2024_COMPLETE_*.xlsx (final output)")
        print("=" * 60 + "\n")
        
        df = scrape_transfermarkt_caf_cup_resumable()
        
        if df is not None:
            print("\n" + "=" * 60)
            print("✓ SCRAPING COMPLETED SUCCESSFULLY!")
            print("=" * 60 + "\n")
        else:
            print("\n" + "=" * 60)
            print("⚠ SCRAPING PAUSED")
            print("=" * 60)
            print("Run the script again to resume from where it stopped.\n")
            
    except KeyboardInterrupt:
        print("\n\n⚠ Scraping interrupted by user.")
        print("Progress has been saved. Run again to resume.\n")
    except Exception as e:
        print(f"\n✗ An unexpected error occurred: {e}\n")
        import traceback
        traceback.print_exc()